# from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem
from PyQt5.QtWidgets import *
import sys
import openpyxl

class Main(QWidget):
  def __init__(self):
    super(Main, self).__init__()
    self.setWindowTitle("Load Excel data to QTableWidget")

    layout = QVBoxLayout()
    self.setLayout(layout)

    self.table_widget = QTableWidget()
    layout.addWidget(self.table_widget)

    ## Load all Excel data
    self.load_data()


  def load_data(self):
    path = r'C:\Users\nicol\OneDrive\Área de Trabalho\TableWidget\list-countries-world.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active ## The currently open sheet

    list_values = list(sheet.values)

    # for value in values:
    #   print(value)

    ## In PyQt5, the table widget always needs the row count and column count to be populated

    self.table_widget.setRowCount(sheet.max_row)
    self.table_widget.setColumnCount(sheet.max_column)

    self.table_widget.setHorizontalHeaderLabels(list_values[0])

    ## Running a for loop and skipping out the first value (slicing [1:]), which is the header

    row_index = 0

    for value_tuple in list_values[1:]:
      print(value_tuple)
      col_index = 0

      for value in value_tuple:
        self.table_widget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
        col_index += 1

      row_index += 1


    ## Inserting any value in a table
    # self.table_widget.setItem(0, 2, QTableWidgetItem("hello")) ## Here we insert the word "hello" in the first row and the third column

if __name__ == "__main__":
  app = QApplication(sys.argv)
  window = Main()
  window.showMaximized()
  app.exec_()

